#!/usr/bin/env python3
"""从 Excel 提取 64强 第一轮数据，写入 archive-site JSON"""

import openpyxl
import json
import os

CN_NUMS = {
    '一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
    '六': 6, '七': 7, '八': 8, '九': 9, '十': 10,
    '十一': 11, '十二': 12, '十三': 13, '十四': 14, '十五': 15,
    '十六': 16, '十七': 17, '十八': 18, '十九': 19, '二十': 20,
    '二十一': 21, '二十二': 22, '二十三': 23, '二十四': 24, '二十五': 25,
    '二十六': 26, '二十七': 27, '二十八': 28, '二十九': 29, '三十': 30,
    '三十一': 31, '三十二': 32, '三十三': 33, '三十四': 34, '三十五': 35,
}

TABLE_LABELS = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1']


def parse_sheet_name(name):
    """从 '联赛杯第X届' 提取数字"""
    if not name.startswith('联赛杯第') or not name.endswith('届'):
        return None
    cn = name[4:-1]
    return CN_NUMS.get(cn)


def extract_table_data(ws, table_num):
    """提取第 table_num 桌的数据 (1-8)"""
    start_row = (table_num - 1) * 11 + 1
    players = []

    for i in range(8):
        row = start_row + 2 + i
        name = ws.cell(row=row, column=1).value
        if not name or isinstance(name, (int, float)):
            continue
        name = str(name).strip()
        if not name or name == 'none':
            continue

        # 判断是否全是 0（没有实际数据）
        has_data = False
        games = []
        placements = []
        for game in range(3):
            col_d = 4 + game * 2  # D=4, F=6, H=8
            col_e = 5 + game * 2  # E=5, G=7, I=9
            placement = ws.cell(row=row, column=col_d).value
            score = ws.cell(row=row, column=col_e).value
            if placement is not None and score is not None:
                try:
                    p = int(placement)
                    s = int(score)
                    if p > 0 and s > 0:
                        placements.append(p)
                        games.append(s)
                        has_data = True
                except (ValueError, TypeError):
                    pass

        if not has_data:
            continue

        total_points = sum(games)
        players.append({
            'name': name,
            'displayName': name,
            'battleTag': name,
            'totalPoints': total_points,
            'games': games,
            'placements': placements,
            'empty': False
        })

    return players


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base, '..', '..', '联赛杯.xlsx')
    archive_dir = os.path.join(base, 'data')

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    updated = []
    skipped = []

    for sheet_name in wb.sheetnames:
        num = parse_sheet_name(sheet_name)
        if num is None or num < 14:
            continue

        ws = wb[sheet_name]

        # 提取 8 桌数据
        groups = []
        for table_num in range(1, 9):
            players = extract_table_data(ws, table_num)
            if not players:
                continue
            max_games = max(len(p['games']) for p in players)
            groups.append({
                'label': TABLE_LABELS[table_num - 1],
                'groupIndex': table_num,
                'boN': max(3, max_games),  # 至少 bo3
                'gamesPlayed': max_games,
                'players': players,
                'nextRoundGroupId': (table_num + 1) // 2
            })

        if not groups:
            skipped.append(f'{sheet_name}: 无有效数据')
            continue

        # 读取已有 JSON
        json_path = os.path.join(archive_dir, f'{sheet_name}.json')
        if not os.path.exists(json_path):
            skipped.append(f'{sheet_name}: JSON 文件不存在')
            continue

        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # 检查是否已有 64强
        for t in data['tournaments']:
            if t['name'] == sheet_name:
                if t['rounds'] and t['rounds'][0]['label'] == '64强':
                    skipped.append(f'{sheet_name}: 已有 64强，跳过')
                    continue

                # 插入 64强 作为第一轮
                new_round = {
                    'label': '64强',
                    'groups': groups
                }
                t['rounds'].insert(0, new_round)

        # 写回 JSON
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        total_players = sum(len(g['players']) for g in groups)
        updated.append(f'{sheet_name}: {len(groups)}桌 {total_players}人')

    # 更新 index.json
    index_path = os.path.join(archive_dir, 'index.json')
    with open(index_path, 'r', encoding='utf-8') as f:
        index = json.load(f)

    for entry in index['tournaments']:
        name = entry['name']
        num = parse_sheet_name(name)
        if num is not None and num >= 14:
            if not any(r['label'] == '64强' for r in entry['rounds']):
                # 检查对应 JSON 是否真的有 64强
                json_path = os.path.join(archive_dir, f'{name}.json')
                if os.path.exists(json_path):
                    with open(json_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    for t in data['tournaments']:
                        if t['name'] == name and t['rounds'] and t['rounds'][0]['label'] == '64强':
                            entry['rounds'].insert(0, {
                                'label': '64强',
                                'groupCount': len(t['rounds'][0]['groups'])
                            })

    with open(index_path, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

    print('=== 已更新 ===')
    for s in updated:
        print(f'  ✅ {s}')
    print(f'\n=== 跳过 ===')
    for s in skipped:
        print(f'  ⏭️ {s}')


if __name__ == '__main__':
    main()
