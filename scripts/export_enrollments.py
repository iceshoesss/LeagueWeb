#!/usr/bin/env python3
"""导出报名名单

用法:
  python export_enrollments.py --csv                   # 导出 CSV（不含 Lo）
  python export_enrollments.py --csv --with-lo         # 导出 CSV（含 Lo，从 league_players/player_records 查询）
  python export_enrollments.py --excel                 # 导出 Excel（不含 Lo）
  python export_enrollments.py --excel --with-lo       # 导出 Excel（含 Lo）
  python export_enrollments.py --missing-lo            # 查找 Lo 为空的玩家

环境变量:
  MONGO_URL  MongoDB 地址 (默认 mongodb://mongo:27017)
  DB_NAME    数据库名 (默认 hearthstone)
"""

import os
import sys
import csv
from datetime import datetime

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://mongo:27017")
DB_NAME = os.environ.get("DB_NAME", "hearthstone")


def get_enrollments():
    from pymongo import MongoClient
    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]
    return list(db.tournament_enrollments.find(
        {"status": {"$in": ["enrolled", "waitlist"]}},
        {"_id": 0, "battleTag": 1, "accountIdLo": 1, "status": 1, "position": 1, "enrollAt": 1}
    ).sort("position", 1))


def lookup_lo(db, battle_tag):
    """从 league_players 和 player_records 查 accountIdLo"""
    lp = db.league_players.find_one({"battleTag": battle_tag}, {"accountIdLo": 1})
    if lp and lp.get("accountIdLo"):
        return str(lp["accountIdLo"])
    pr = db.player_records.find_one({"playerId": battle_tag}, {"accountIdLo": 1})
    if pr and pr.get("accountIdLo"):
        return str(pr["accountIdLo"])
    return ""


def check_missing_lo():
    """查找 Lo 为空的玩家，交叉比对 league_players 和 player_records"""
    from pymongo import MongoClient
    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]

    enrollments = get_enrollments()
    if not enrollments:
        print("暂无报名数据")
        return

    missing = []
    for r in enrollments:
        lo = str(r.get("accountIdLo", "")).strip()
        if not lo or lo == "None":
            missing.append(r)

    if not missing:
        print(f"✅ 全部 {len(enrollments)} 人都有 Lo，没有问题")
        return

    print(f"⚠️  共 {len(missing)} 人 Lo 为空（总计 {len(enrollments)} 人报名）\n")
    print(f"{'序号':>4}  {'BattleTag':<30}  {'league_players Lo':<18}  {'player_records Lo':<18}  {'状态'}")
    print("-" * 100)

    for i, r in enumerate(missing, 1):
        tag = r.get("battleTag", "")
        status = "替补" if r.get("status") == "waitlist" else "正选"

        lp = db.league_players.find_one({"battleTag": tag}, {"accountIdLo": 1})
        lp_lo = str(lp.get("accountIdLo", "")) if lp else "-"

        pr = db.player_records.find_one({"playerId": tag}, {"accountIdLo": 1})
        pr_lo = str(pr.get("accountIdLo", "")) if pr else "-"

        print(f"{i:>4}  {tag:<30}  {lp_lo:<18}  {pr_lo:<18}  {status}")

    print()
    print("修复方法:")
    print("  1. 玩家用 bg_tool/HDT 插件打一局，插件会自动上报 accountIdLo")
    print("  2. 管理员手动补 Lo: db.tournament_enrollments.updateOne({battleTag: 'xxx'}, {$set: {accountIdLo: '12345'}})")
    print("  3. 同步到 league_players: db.league_players.updateOne({battleTag: 'xxx'}, {$set: {accountIdLo: '12345'}})")


def export_csv(rows, with_lo):
    from pymongo import MongoClient
    db = MongoClient(MONGO_URL)[DB_NAME] if with_lo else None

    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{ts}_报名名单{'_含Lo' if with_lo else ''}.csv"

    if with_lo:
        headers = ["序号", "BattleTag", "AccountIdLo", "报名时间", "状态"]
    else:
        headers = ["序号", "BattleTag", "报名时间", "状态"]

    with open(filename, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for i, r in enumerate(rows, 1):
            status = "替补" if r.get("status") == "waitlist" else ""
            if with_lo:
                lo = lookup_lo(db, r.get("battleTag", ""))
                writer.writerow([i, r.get("battleTag", ""), lo, r.get("enrollAt", ""), status])
            else:
                writer.writerow([i, r.get("battleTag", ""), r.get("enrollAt", ""), status])

    print(f"已导出: {filename} ({len(rows)} 人)")


def export_excel(rows, with_lo):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Excel 模式需要 openpyxl: pip install openpyxl")
        sys.exit(1)

    from pymongo import MongoClient
    db = MongoClient(MONGO_URL)[DB_NAME] if with_lo else None

    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{ts}_报名名单{'_含Lo' if with_lo else ''}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "报名名单"

    if with_lo:
        headers = ["序号", "BattleTag", "AccountIdLo", "报名时间", "状态"]
    else:
        headers = ["序号", "BattleTag", "报名时间", "状态"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2a2a4a", end_color="2a2a4a", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="cccccc"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, 2):
        status = "替补" if r.get("status") == "waitlist" else ""
        ws.cell(row=i, column=1, value=i - 1).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=r.get("battleTag", ""))

        if with_lo:
            lo = lookup_lo(db, r.get("battleTag", ""))
            ws.cell(row=i, column=3, value=lo)
            ws.cell(row=i, column=4, value=r.get("enrollAt", ""))
            cell_status = ws.cell(row=i, column=5, value=status)
            col_count = 5
        else:
            ws.cell(row=i, column=3, value=r.get("enrollAt", ""))
            cell_status = ws.cell(row=i, column=4, value=status)
            col_count = 4

        if status == "替补":
            cell_status.font = Font(color="e2b714")
        for col in range(1, col_count + 1):
            ws.cell(row=i, column=col).border = thin_border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    if with_lo:
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 8
    else:
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 8

    wb.save(filename)
    print(f"已导出: {filename} ({len(rows)} 人)")


def main():
    args = sys.argv[1:]
    with_lo = "--with-lo" in args
    args = [a for a in args if a != "--with-lo"]

    if not args or args[0] not in ("--csv", "--excel", "--missing-lo"):
        print("用法:")
        print("  python export_enrollments.py --csv [--with-lo]")
        print("  python export_enrollments.py --excel [--with-lo]")
        print("  python export_enrollments.py --missing-lo")
        sys.exit(1)

    if args[0] == "--missing-lo":
        check_missing_lo()
        return

    rows = get_enrollments()
    if not rows:
        print("暂无报名数据")
        return

    if args[0] == "--csv":
        export_csv(rows, with_lo)
    else:
        export_excel(rows, with_lo)


if __name__ == "__main__":
    main()
