#!/usr/bin/env python3
"""导出晋级名单

用法:
  python export_qualifiers.py <赛事名称> --csv     # 导出 CSV
  python export_qualifiers.py <赛事名称> --excel   # 导出 Excel

示例:
  python export_qualifiers.py "2026 春季赛" --csv
"""

import os
import sys
import csv
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from data import get_group_rankings

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://mongo:27017")
DB_NAME = os.environ.get("DB_NAME", "hearthstone")


def get_qualifiers(tournament_name):
    from pymongo import MongoClient
    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]

    groups = list(db.tournament_groups.find({
        "tournamentName": tournament_name,
        "status": "done",
    }).sort([("round", 1), ("groupIndex", 1)]))

    if not groups:
        print(f"赛事「{tournament_name}」没有已完成的分组")
        return []

    rankings = get_group_rankings(db, tournament_name)

    result = []
    for g in groups:
        tg_str = str(g["_id"])
        rd = g.get("round", 0)
        gi = g.get("groupIndex", 0)
        group_rankings = rankings.get(tg_str, {})

        ranked = sorted(
            g.get("players", []),
            key=lambda p: group_rankings.get(str(p.get("accountIdLo", "")), {}).get("totalPoints", 0),
            reverse=True,
        )

        for i, p in enumerate(ranked):
            lo = str(p.get("accountIdLo", ""))
            rd_data = group_rankings.get(lo, {})
            result.append({
                "round": rd,
                "groupIndex": gi,
                "rank": i + 1,
                "qualified": i < 4,
                "battleTag": p.get("battleTag", ""),
                "displayName": p.get("displayName", ""),
                "accountIdLo": lo,
                "heroName": p.get("heroName", ""),
                "totalPoints": rd_data.get("totalPoints", 0),
                "games": rd_data.get("games", []),
                "chickens": rd_data.get("chickens", 0),
            })

    return result


def export_csv(rows, tournament_name):
    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{ts}_晋级名单.csv"
    with open(filename, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["轮次", "组号", "组内排名", "晋级", "BattleTag", "显示名", "总积分", "各局得分", "吃鸡"])
        for r in rows:
            writer.writerow([
                r["round"],
                r["groupIndex"],
                r["rank"],
                "✅" if r["qualified"] else "",
                r["battleTag"],
                r["displayName"],
                r["totalPoints"],
                "/".join(str(x) for x in r["games"]),
                r["chickens"],
            ])
    print(f"已导出: {filename} ({len(rows)} 人，其中晋级 {sum(1 for r in rows if r['qualified'])} 人)")


def export_excel(rows, tournament_name):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Excel 模式需要 openpyxl: pip install openpyxl")
        sys.exit(1)

    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{ts}_晋级名单.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "晋级名单"

    headers = ["轮次", "组号", "组内排名", "晋级", "BattleTag", "显示名", "总积分", "各局得分", "吃鸡"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2a2a4a", end_color="2a2a4a", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="cccccc"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r["round"]).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=r["groupIndex"]).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3, value=r["rank"]).alignment = Alignment(horizontal="center")
        qual_cell = ws.cell(row=i, column=4, value="✅" if r["qualified"] else "")
        qual_cell.alignment = Alignment(horizontal="center")
        if r["qualified"]:
            qual_cell.font = Font(color="22c55e")
        ws.cell(row=i, column=5, value=r["battleTag"])
        ws.cell(row=i, column=6, value=r["displayName"])
        ws.cell(row=i, column=7, value=r["totalPoints"]).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=8, value="/".join(str(x) for x in r["games"]))
        ws.cell(row=i, column=9, value=r["chickens"]).alignment = Alignment(horizontal="center")
        for col in range(1, 10):
            ws.cell(row=i, column=col).border = thin_border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 6

    wb.save(filename)
    print(f"已导出: {filename} ({len(rows)} 人，其中晋级 {sum(1 for r in rows if r['qualified'])} 人)")


def main():
    if len(sys.argv) < 3:
        print("用法: python export_qualifiers.py <赛事名称> --csv | --excel")
        sys.exit(1)

    tournament_name = sys.argv[1]
    fmt = sys.argv[2]

    if fmt not in ("--csv", "--excel"):
        print("用法: python export_qualifiers.py <赛事名称> --csv | --excel")
        sys.exit(1)

    rows = get_qualifiers(tournament_name)
    if not rows:
        return

    if fmt == "--csv":
        export_csv(rows, tournament_name)
    else:
        export_excel(rows, tournament_name)


if __name__ == "__main__":
    main()
