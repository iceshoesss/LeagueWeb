#!/usr/bin/env python3
"""导出晋级 + 种子选手

根据赛事名称，导出晋级选手（前 4 名）和种子选手。
晋级选手在上，种子选手在下（Excel 中种子行标浅绿色）。

用法:
  python export_qualifiers_seeds.py <赛事名称> --csv
  python export_qualifiers_seeds.py <赛事名称> --excel
  python export_qualifiers_seeds.py <赛事名称> --csv --with-lo
  python export_qualifiers_seeds.py <赛事名称> --excel --with-lo

示例:
  python export_qualifiers_seeds.py "2026 春季赛" --csv
  python export_qualifiers_seeds.py "2026 春季赛" --excel --with-lo
"""

import os
import sys
import csv
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from data import get_group_rankings

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://mongo:27017")
DB_NAME = os.environ.get("DB_NAME", "hearthstone")


def get_qualifiers(db, tournament_name):
    groups = list(db.tournament_groups.find({
        "tournamentName": tournament_name,
        "status": "done",
    }).sort([("round", 1), ("groupIndex", 1)]))

    if not groups:
        return []

    rankings = get_group_rankings(db, tournament_name)
    seen = set()
    result = []
    for g in groups:
        tg_str = str(g["_id"])
        group_rankings = rankings.get(tg_str, {})
        ranked = sorted(
            g.get("players", []),
            key=lambda p: group_rankings.get(str(p.get("accountIdLo", "")), {}).get("totalPoints", 0),
            reverse=True,
        )
        for i, p in enumerate(ranked):
            if i >= 4:
                break
            lo = str(p.get("accountIdLo", ""))
            tag = p.get("battleTag", "")
            if tag and tag not in seen:
                seen.add(tag)
                rd_data = group_rankings.get(lo, {})
                result.append({
                    "battleTag": tag,
                    "accountIdLo": lo,
                    "totalPoints": rd_data.get("totalPoints", 0),
                })
    return result


def get_seeds(db):
    return list(db.league_players.find(
        {"isSeed": True},
        {"_id": 0, "battleTag": 1, "accountIdLo": 1}
    ).sort("battleTag", 1))


def export_csv(qualifiers, seeds, with_lo):
    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{ts}_晋级+种子{'_含Lo' if with_lo else ''}.csv"

    if with_lo:
        headers = ["序号", "BattleTag", "AccountIdLo", "类型"]
    else:
        headers = ["序号", "BattleTag", "类型"]

    with open(filename, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        idx = 0
        for p in qualifiers:
            idx += 1
            tag = p["battleTag"]
            if with_lo:
                writer.writerow([idx, tag, p["accountIdLo"], "晋级"])
            else:
                writer.writerow([idx, tag, "晋级"])
        for p in seeds:
            idx += 1
            tag = p.get("battleTag", "")
            if with_lo:
                writer.writerow([idx, tag, str(p.get("accountIdLo", "")), "种子"])
            else:
                writer.writerow([idx, tag, "种子"])

    print(f"已导出: {filename} (晋级 {len(qualifiers)} 人 + 种子 {len(seeds)} 人)")


def export_excel(qualifiers, seeds, with_lo):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Excel 模式需要 openpyxl: pip install openpyxl")
        sys.exit(1)

    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{ts}_晋级+种子{'_含Lo' if with_lo else ''}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "晋级+种子"

    if with_lo:
        headers = ["序号", "BattleTag", "AccountIdLo", "类型"]
    else:
        headers = ["序号", "BattleTag", "类型"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2a2a4a", end_color="2a2a4a", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="cccccc"))
    seed_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")  # 浅绿

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    idx = 0

    # 晋级选手
    for p in qualifiers:
        idx += 1
        tag = p["battleTag"]
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=tag)
        if with_lo:
            ws.cell(row=row, column=3, value=p["accountIdLo"])
            ws.cell(row=row, column=4, value="晋级").alignment = Alignment(horizontal="center")
            col_count = 4
        else:
            ws.cell(row=row, column=3, value="晋级").alignment = Alignment(horizontal="center")
            col_count = 3
        for c in range(1, col_count + 1):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    # 种子选手（浅绿底色）
    for p in seeds:
        idx += 1
        tag = p.get("battleTag", "")
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).fill = seed_fill
        ws.cell(row=row, column=2, value=tag)
        ws.cell(row=row, column=2).fill = seed_fill
        if with_lo:
            ws.cell(row=row, column=3, value=str(p.get("accountIdLo", "")))
            ws.cell(row=row, column=3).fill = seed_fill
            ws.cell(row=row, column=4, value="种子").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4).fill = seed_fill
            col_count = 4
        else:
            ws.cell(row=row, column=3, value="种子").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3).fill = seed_fill
            col_count = 3
        for c in range(1, col_count + 1):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    if with_lo:
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 8
    else:
        ws.column_dimensions["C"].width = 8

    wb.save(filename)
    print(f"已导出: {filename} (晋级 {len(qualifiers)} 人 + 种子 {len(seeds)} 人)")


def main():
    args = sys.argv[1:]
    with_lo = "--with-lo" in args
    args = [a for a in args if a != "--with-lo"]

    if len(args) < 2 or args[-1] not in ("--csv", "--excel"):
        print("用法:")
        print("  python export_qualifiers_seeds.py <赛事名称> --csv [--with-lo]")
        print("  python export_qualifiers_seeds.py <赛事名称> --excel [--with-lo]")
        sys.exit(1)

    tournament_name = args[0]
    fmt = args[-1]

    from pymongo import MongoClient
    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]

    qualifiers = get_qualifiers(db, tournament_name)
    seeds = get_seeds(db)

    if not qualifiers and not seeds:
        print("暂无数据")
        return

    print(f"晋级 {len(qualifiers)} 人 + 种子 {len(seeds)} 人：")
    for p in qualifiers:
        print(f"  [晋级] {p['battleTag']}")
    for p in seeds:
        print(f"  [种子] {p.get('battleTag', '?')}")
    print()

    if fmt == "--csv":
        export_csv(qualifiers, seeds, with_lo)
    else:
        export_excel(qualifiers, seeds, with_lo)


if __name__ == "__main__":
    main()
