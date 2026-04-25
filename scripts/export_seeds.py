#!/usr/bin/env python3
"""导出种子选手

从 league_players 导出 isSeed=True 的选手。

用法:
  python export_seeds.py --csv
  python export_seeds.py --excel
  python export_seeds.py --csv --with-lo
  python export_seeds.py --excel --with-lo

环境变量:
  MONGO_URL  MongoDB 地址 (默认 mongodb://mongo:27017)
  DB_NAME    数据库名 (默认 hearthstone)
"""

import os
import sys
import csv
from datetime import datetime

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://mongo:27017")
DB_NAME = os.environ.get("DB_NAME", "hearthstone")


def get_seeds():
    from pymongo import MongoClient
    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]
    return db, list(db.league_players.find(
        {"isSeed": True},
        {"_id": 0, "battleTag": 1, "displayName": 1, "accountIdLo": 1}
    ).sort("battleTag", 1))


def export_csv(seeds, with_lo):
    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{ts}_种子选手{'_含Lo' if with_lo else ''}.csv"

    if with_lo:
        headers = ["序号", "BattleTag", "AccountIdLo"]
    else:
        headers = ["序号", "BattleTag"]

    with open(filename, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for i, p in enumerate(seeds, 1):
            tag = p.get("battleTag", "")
            if with_lo:
                lo = str(p.get("accountIdLo", ""))
                writer.writerow([i, tag, lo])
            else:
                writer.writerow([i, tag])

    print(f"已导出: {filename} ({len(seeds)} 人)")


def export_excel(seeds, with_lo):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Excel 模式需要 openpyxl: pip install openpyxl")
        sys.exit(1)

    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{ts}_种子选手{'_含Lo' if with_lo else ''}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "种子选手"

    if with_lo:
        headers = ["序号", "BattleTag", "AccountIdLo"]
    else:
        headers = ["序号", "BattleTag"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2a2a4a", end_color="2a2a4a", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="cccccc"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, p in enumerate(seeds, 2):
        tag = p.get("battleTag", "")
        ws.cell(row=i, column=1, value=i - 1).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=tag)
        if with_lo:
            lo = str(p.get("accountIdLo", ""))
            ws.cell(row=i, column=3, value=lo)
            col_count = 3
        else:
            col_count = 2
        for c in range(1, col_count + 1):
            ws.cell(row=i, column=c).border = thin_border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    if with_lo:
        ws.column_dimensions["C"].width = 16

    wb.save(filename)
    print(f"已导出: {filename} ({len(seeds)} 人)")


def main():
    args = sys.argv[1:]
    with_lo = "--with-lo" in args
    args = [a for a in args if a != "--with-lo"]

    if not args or args[0] not in ("--csv", "--excel"):
        print("用法:")
        print("  python export_seeds.py --csv [--with-lo]")
        print("  python export_seeds.py --excel [--with-lo]")
        sys.exit(1)

    _, seeds = get_seeds()
    if not seeds:
        print("暂无种子选手")
        return

    print(f"共 {len(seeds)} 位种子选手：")
    for p in seeds:
        print(f"  {p.get('battleTag', '?')}")
    print()

    if args[0] == "--csv":
        export_csv(seeds, with_lo)
    else:
        export_excel(seeds, with_lo)


if __name__ == "__main__":
    main()
