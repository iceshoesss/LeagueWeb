#!/usr/bin/env python3
"""导出海选分组

根据赛事名称，从 tournament_groups 导出分组数据。
组名循环：A1-D1、A2-D2、A3-D3 ...

用法:
  python export_groups.py <赛事名称> --csv
  python export_groups.py <赛事名称> --excel
  python export_groups.py <赛事名称> --csv --with-lo
  python export_groups.py <赛事名称> --excel --with-lo

示例:
  python export_groups.py "2026 春季赛" --csv
  python export_groups.py "2026 春季赛" --excel --with-lo
"""

import os
import sys
import csv
from datetime import datetime

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://mongo:27017")
DB_NAME = os.environ.get("DB_NAME", "hearthstone")

GROUP_LABELS = "ABCD"


def group_label(group_index):
    """0-based group_index → A1, B1, C1, D1, A2, B2, ..."""
    return f"{GROUP_LABELS[group_index % 4]}{group_index // 4 + 1}"


def lookup_lo(db, battle_tag):
    """从 league_players 和 player_records 查 accountIdLo"""
    lp = db.league_players.find_one({"battleTag": battle_tag}, {"accountIdLo": 1})
    if lp and lp.get("accountIdLo"):
        return str(lp["accountIdLo"])
    pr = db.player_records.find_one({"playerId": battle_tag}, {"accountIdLo": 1})
    if pr and pr.get("accountIdLo"):
        return str(pr["accountIdLo"])
    return ""


def get_groups(tournament_name):
    from pymongo import MongoClient
    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]

    groups = list(db.tournament_groups.find({
        "tournamentName": tournament_name,
        "round": 1,
    }).sort("groupIndex", 1))

    if not groups:
        print(f"赛事「{tournament_name}」没有分组数据")
        return None, []

    return db, groups


def export_csv(db, groups, tournament_name, with_lo):
    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{ts}_{tournament_name}_分组{'_含Lo' if with_lo else ''}.csv"

    if with_lo:
        headers = ["组别", "序号", "BattleTag", "显示名", "AccountIdLo"]
    else:
        headers = ["组别", "序号", "BattleTag", "显示名"]

    total = 0
    with open(filename, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for gi, g in enumerate(groups):
            label = group_label(g.get("groupIndex", 1) - 1)
            for i, p in enumerate(g.get("players", []), 1):
                if p.get("empty"):
                    continue
                tag = p.get("battleTag", "")
                display = p.get("displayName", "")
                if with_lo:
                    lo = p.get("accountIdLo", "") or lookup_lo(db, tag)
                    writer.writerow([label, i, tag, display, lo])
                else:
                    writer.writerow([label, i, tag, display])
                total += 1

    print(f"已导出: {filename} ({total} 人，{len(groups)} 组)")


def export_excel(db, groups, tournament_name, with_lo):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Excel 模式需要 openpyxl: pip install openpyxl")
        sys.exit(1)

    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{ts}_{tournament_name}_分组{'_含Lo' if with_lo else ''}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "海选分组"

    if with_lo:
        headers = ["组别", "序号", "BattleTag", "显示名", "AccountIdLo"]
    else:
        headers = ["组别", "序号", "BattleTag", "显示名"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2a2a4a", end_color="2a2a4a", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="cccccc"))
    group_fills = [
        PatternFill(start_color="f0f4ff", end_color="f0f4ff", fill_type="solid"),  # A - 浅蓝
        PatternFill(start_color="f0fff4", end_color="f0fff4", fill_type="solid"),  # B - 浅绿
        PatternFill(start_color="fff8f0", end_color="fff8f0", fill_type="solid"),  # C - 浅橙
        PatternFill(start_color="f8f0ff", end_color="f8f0ff", fill_type="solid"),  # D - 浅紫
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    total = 0
    for gi, g in enumerate(groups):
        label = group_label(g.get("groupIndex", 1) - 1)
        group_idx = "ABCD".index(label[0]) if label[0] in "ABCD" else 0
        fill = group_fills[group_idx % 4]

        for i, p in enumerate(g.get("players", []), 1):
            if p.get("empty"):
                continue
            tag = p.get("battleTag", "")
            display = p.get("displayName", "")

            ws.cell(row=row, column=1, value=label).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=2, value=i).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2).fill = fill
            ws.cell(row=row, column=3, value=tag)
            ws.cell(row=row, column=3).fill = fill
            ws.cell(row=row, column=4, value=display)
            ws.cell(row=row, column=4).fill = fill

            if with_lo:
                lo = p.get("accountIdLo", "") or lookup_lo(db, tag)
                ws.cell(row=row, column=5, value=lo)
                ws.cell(row=row, column=5).fill = fill
                col_count = 5
            else:
                col_count = 4

            for c in range(1, col_count + 1):
                ws.cell(row=row, column=c).border = thin_border
            row += 1
            total += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 20
    if with_lo:
        ws.column_dimensions["E"].width = 16

    wb.save(filename)
    print(f"已导出: {filename} ({total} 人，{len(groups)} 组)")


def main():
    args = sys.argv[1:]
    with_lo = "--with-lo" in args
    args = [a for a in args if a != "--with-lo"]

    if len(args) < 2 or args[-1] not in ("--csv", "--excel"):
        print("用法:")
        print("  python export_groups.py <赛事名称> --csv [--with-lo]")
        print("  python export_groups.py <赛事名称> --excel [--with-lo]")
        sys.exit(1)

    tournament_name = args[0]
    fmt = args[-1]

    db, groups = get_groups(tournament_name)
    if not groups:
        return

    print(f"赛事「{tournament_name}」共 {len(groups)} 组：")
    for g in groups:
        label = group_label(g.get("groupIndex", 1) - 1)
        names = [p.get("displayName", "?") for p in g.get("players", []) if not p.get("empty")]
        print(f"  {label} 组 ({len(names)} 人): {', '.join(names)}")
    print()

    if fmt == "--csv":
        export_csv(db, groups, tournament_name, with_lo)
    else:
        export_excel(db, groups, tournament_name, with_lo)


if __name__ == "__main__":
    main()
